import time
import pandas as pd
import csv
from instaloader import Instaloader, Profile, exceptions
from datetime import datetime, timedelta
from itertools import dropwhile, takewhile
from os import listdir, mkdir
import openpyxl

archivo_excel_lectura = 'C:\\NAS_PAU\\TFG\\PGA\\PGACompletoDepurado.xlsx'

def descarga_posts_entre_fechas(str_profile, dt_fecha_inicio, dt_fecha_final, str_directorio):
    """
    Guarda los posts entre las fechas indicadas (ambas 
    inclusive) y genera un archivo resumen en formato csv
    """

    # Generar instancia de Instaloader
    L = Instaloader(compress_json=False, save_metadata=False)
    L.load_session_from_file('jugadoresgolf22')
    profile = Profile.from_username(L.context, str_profile)  # Descargar el perfil
    posts = profile.get_posts()  # Recupera todos los posts del perfil cargado

    # Campos del csv
    campos_csv = ['num', 'post', 'shortcode', 'fecha', 'hora', 'video', 'likes', 'comentarios', 'texto', 'hashtags'] 
    lista_filas = []

    # Contador de posts
    n = 0

    # Descarga iterativamente los posts del perfil dentro del intervalo de fechas indicado
    for post in takewhile(lambda p: p.date > dt_fecha_inicio, dropwhile(lambda p: p.date > dt_fecha_final + timedelta(1), posts)):
        try:
            L.download_post(post, str_directorio)  # Intentar descargar el post en el directorio
        except Instaloader.exceptions.InstaloaderException as e:
            print(f"Error al descargar la imagen: {post}")  # Si hay error en la descarga

        n += 1  # Contador para saber cuantos posts ha publicado el usuario en el intervalo de fechas

        # Generar el archivo resumen
        fila = {}
        fila['num'] = 0  # Inicializar el índice de la fila
        fila['post'] = post.date.strftime('%Y-%m-%d_%H-%M-%S_UTC')
        fila['shortcode'] = post.shortcode
        dt_post_corregido = post.date + timedelta(hours=1)  # Ajustamos la hora de UTC a horario de España
        fila['fecha'] = dt_post_corregido.strftime('%Y-%m-%d')
        fila['hora'] = dt_post_corregido.strftime('%H:%M:%S')
        fila['video'] = 1 if post.is_video else 0
        fila['likes'] = post.likes
        fila['comentarios'] = post.comments
        if post.caption is not None:
            fila['texto'] = post.caption.replace('\n', ' ').replace('\r', ' ')
            fila['hashtags'] = '; '.join(list(set(post.caption_hashtags)))
        
        lista_filas.append(fila)

        # Esto es para no saturar el servidor y evitar la desconexión
        if n % 50 == 0:
            print('Pausa para evitar bloqueos rápidos')
            time.sleep(120)  # Pausa más larga para evitar bloqueos rápidos

    # Actualizar el índice de los posts
    for n_fila, fila in enumerate(reversed(lista_filas)):
        fila['num'] = n_fila + 1

    # Guardamos el resumen de los posts descargados del usuario entre las fechas indicadas en un archivo csv
    with open(str_directorio + '/0_' + str_profile + '_resumen.csv', 'w', encoding='UTF-8') as archivo:
        writer = csv.DictWriter(archivo, fieldnames=campos_csv, delimiter='\t', lineterminator='\n')
        writer.writeheader()
        writer.writerows(reversed(lista_filas))

    return n

def seguidores(str_profile):
    """Devuelve el número de seguidores de un perfil de Instagram"""
    L = Instaloader()
    profile = Profile.from_username(L.context, str_profile)
    return profile.followers

def extraer_username(instagram_url):
    """Extrae el username desde una URL o devuelve el username si ya lo es"""
    if instagram_url.startswith('http'):
        return instagram_url.rstrip('/').split('/')[-1]
    return instagram_url

# Función de reintento en caso de 401 Unauthorized
def manejar_error_401(func, *args, **kwargs):
    """Reintenta la función después de un error 401 Unauthorized"""
    while True:
        try:
            return func(*args, **kwargs)
        except exceptions.InstaloaderException as e:
            if "401 Unauthorized" in str(e):
                print("Error 401 detectado. Esperando 10 minutos antes de reintentar...")
                time.sleep(600)  # Espera 10 minutos
            else:
                raise e  # Si el error no es 401, lo volvemos a lanzar

# Archivo con los usuarios a descargar
dataframe = pd.read_excel(archivo_excel_lectura)

users = []

# Creamos la lista con los usuarios de Instagram pendientes de descargar 
# (Esto se ejecuta después de que haya dado error por desconexión del servidor)
for indice_fila, fila in dataframe.iterrows():
    user = fila['Instagram']
    descargado = fila['Descargado']

    if descargado == 0:
        users.append(user)

len(users)

# Abrimos el archivo Excel con los usuarios de Instagram,
# modificamos la columna "DESCARGADO" y añadimos el número de posts y seguidores
wb = openpyxl.load_workbook(archivo_excel)
hoja = wb.active

for user in users:
    str_profile = extraer_username(str(user))

    # Periodo a descargar (en este caso el año entre las dos últimas ediciones de los Premios Grammy)
    inicio = datetime(2024, 1, 1)
    final = datetime(2024, 12, 31)
    if str_profile not in listdir('.'):
        mkdir(str_profile)

    print(f'Descargando posts en el directorio {str_profile}...')
    
    # Usamos la función que maneja el error 401
    n_posts = manejar_error_401(descarga_posts_entre_fechas, str_profile, inicio, final, str_profile)
    followers = seguidores(str_profile)
    print(f'OPERACIÓN FINALIZADA\n - Se han descargado {n_posts} posts de la cuenta de Instagram de {str_profile}')
    
    # Actualizamos los datos en el archivo Excel
    for index, row in enumerate(hoja.iter_rows(min_row=1, values_only=True)):
        if row[1] == str_profile:
            hoja.cell(row=index + 1, column=17, value=1)
            print(f' - La cuenta de Instagram {str_profile} ha sido marcada como descargada')
            hoja.cell(row=index + 1, column=18, value=n_posts)
            hoja.cell(row=index + 1, column=19, value=followers)
            print(f'- La cuenta de Instagram {str_profile} tiene {followers} seguidores\n')

wb.save(archivo_excel)

